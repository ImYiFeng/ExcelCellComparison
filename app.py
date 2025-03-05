from flask import Flask, request, send_file, jsonify, render_template
import pandas as pd
import re
import os
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024


# 解析单元格引用，返回行索引和列索引
def parse_cell_reference(ref: str):
    match = re.match(r"([A-Za-z]+)(\d+)", ref.strip())
    if not match:
        raise ValueError(f"无效的单元格引用：{ref}")

    col_letters = match.group(1).upper()
    row_number = int(match.group(2))

    col_index = 0
    for c in col_letters:
        col_index = col_index * 26 + (ord(c) - ord('A') + 1)
    col_index -= 1

    row_index = row_number - 1

    return row_index, col_index


# 根据分隔符拆分单元格内容，返回拆分后的列表
def split_cell_content(content: str, delimiters: str):
    if not delimiters.strip():
        parts = content.splitlines()
    else:
        escaped_delims = [re.escape(d) for d in delimiters]
        escaped_delims.append(r'\n')
        pattern = "|".join(escaped_delims)
        parts = re.split(pattern, content)

    parts = [p.strip() for p in parts if p.strip() != ""]
    return parts


# 查找列表中的重复项，返回重复项的集合
def find_duplicates(lst):
    seen = set()
    duplicates = set()
    for item in lst:
        if item in seen:
            duplicates.add(item)
        else:
            seen.add(item)
    return duplicates

# 去除列表中的重复项，返回去重后的列表
def deduplicate_list(lst):
    seen = set()
    dedup = []
    for item in lst:
        if item not in seen:
            dedup.append(item)
            seen.add(item)
    return dedup


# 比较两个去重后的列表，返回各自独有的内容
def compare_contents(dedup1, dedup2):
    set1 = set(dedup1)
    set2 = set(dedup2)

    unique1 = list(set1 - set2)
    unique2 = list(set2 - set1)

    return unique1, unique2


# 生成结果Excel文件
def create_result_workbook(cell_ref_1, cell_ref_2, cell1_list, cell2_list,
                           duplicates1, duplicates2, cell1_dedup, cell2_dedup,
                           unique1, unique2):
    wb = Workbook()
    ws = wb.active
    ws.title = "比较结果"

    # 设置表头
    headers = [
        f"单元格 {cell_ref_1} 原始内容",
        f"单元格 {cell_ref_1} 去重后内容",
        f"单元格 {cell_ref_2} 原始内容",
        f"单元格 {cell_ref_2} 去重后内容",
        f"单元格 {cell_ref_1} 独有内容",
        f"单元格 {cell_ref_2} 独有内容"
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # 样式定义
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # 重复项
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 独有项

    # 计算最大行数
    max_rows = max(
        len(cell1_list),
        len(cell2_list),
        len(cell1_dedup),
        len(cell2_dedup),
        len(unique1),
        len(unique2)
    )

    # 填充数据
    for row_idx in range(2, max_rows + 2):  # 从第二行开始
        # 原始内容 - 单元格1
        if row_idx - 2 < len(cell1_list):
            cell = ws.cell(row=row_idx, column=1, value=cell1_list[row_idx - 2])
            if cell.value in duplicates1:
                cell.fill = blue_fill

        # 去重内容 - 单元格1
        if row_idx - 2 < len(cell1_dedup):
            cell = ws.cell(row=row_idx, column=2, value=cell1_dedup[row_idx - 2])
            if cell.value in unique1:
                cell.fill = yellow_fill

        # 原始内容 - 单元格2
        if row_idx - 2 < len(cell2_list):
            cell = ws.cell(row=row_idx, column=3, value=cell2_list[row_idx - 2])
            if cell.value in duplicates2:
                cell.fill = blue_fill

        # 去重内容 - 单元格2
        if row_idx - 2 < len(cell2_dedup):
            cell = ws.cell(row=row_idx, column=4, value=cell2_dedup[row_idx - 2])
            if cell.value in unique2:
                cell.fill = yellow_fill

        # 独有内容 - 单元格1
        if row_idx - 2 < len(unique1):
            ws.cell(row=row_idx, column=5, value=unique1[row_idx - 2])

        # 独有内容 - 单元格2
        if row_idx - 2 < len(unique2):
            ws.cell(row=row_idx, column=6, value=unique2[row_idx - 2])

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母

        # 计算最大内容长度
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except: pass

        # 设置列宽
        ws.column_dimensions[column].width = max_length + 2

    return wb


@app.route('/get-sheets', methods=['POST'])
def get_sheets():
    """获取Excel文件的工作表列表"""
    if 'excel_file' not in request.files:
        return jsonify(error="请先上传Excel文件"), 400

    file = request.files['excel_file']
    try:
        xl = pd.ExcelFile(file)
        return jsonify(sheets=xl.sheet_names)
    except Exception as e:
        return jsonify(error=f"读取Excel失败: {str(e)}"), 400


@app.route('/compare', methods=['POST'])
def handle_comparison():
    try:
        # 验证必需参数
        required_fields = ['sheet_name', 'cell1', 'cell2']
        for field in required_fields:
            if not request.form.get(field):
                return jsonify(error=f"缺少必需参数: {field}"), 400

        # 获取表单数据
        sheet_name = request.form['sheet_name']
        cell_ref_1 = request.form['cell1']
        cell_ref_2 = request.form['cell2']
        delimiters = request.form.get('delimiters', '')

        # 处理文件上传
        if 'excel_file' not in request.files:
            return jsonify(error="请上传Excel文件"), 400

        excel_file = request.files['excel_file']
        if not excel_file.filename.lower().endswith(('.xls', '.xlsx')):
            return jsonify(error="仅支持Excel文件 (.xls, .xlsx)"), 400

        # 读取Excel文件
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        except Exception as e:
            return jsonify(error=f"读取Excel失败: {str(e)}"), 400

        # 解析单元格引用
        try:
            row1, col1 = parse_cell_reference(cell_ref_1)
            row2, col2 = parse_cell_reference(cell_ref_2)
        except ValueError as e:
            return jsonify(error=str(e)), 400

        # 读取单元格内容
        try:
            cell1_value = str(df.iat[row1, col1])
            cell2_value = str(df.iat[row2, col2])
        except IndexError as e:
            return jsonify(error=f"单元格索引越界: {str(e)}"), 400

        # 处理单元格内容
        cell1_list = split_cell_content(cell1_value, delimiters)
        cell2_list = split_cell_content(cell2_value, delimiters)

        duplicates1 = find_duplicates(cell1_list)
        duplicates2 = find_duplicates(cell2_list)

        cell1_dedup = deduplicate_list(cell1_list)
        cell2_dedup = deduplicate_list(cell2_list)

        unique1, unique2 = compare_contents(cell1_dedup, cell2_dedup)

        # 生成结果文件
        wb = create_result_workbook(cell_ref_1, cell_ref_2, cell1_list, cell2_list,
                                    duplicates1, duplicates2, cell1_dedup, cell2_dedup,
                                    unique1, unique2)

        # 保存到临时文件
        with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            result_data = tmp.read()

        os.unlink(tmp.name)  # 删除临时文件

        # 返回文件
        return send_file(
            BytesIO(result_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='comparison_result.xlsx'
        )

    except Exception as e:
        return jsonify(error=f"服务器内部错误: {str(e)}"), 500


@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=9100)
